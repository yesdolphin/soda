def part1():
  import os
  x = """import os
  os.chdir(os.path.expanduser("~/tavern/tavern"))
  from soda.can import *
  def n1():
    x=Shop()
    shop_abbre = OSA.log("input 3 letter shop abbreviation for shop")
    file_loc = OSA.log("input workbook location")
    import openpyxl

    wb = openpyxl.load_workbook(file_loc,data_only=True)
    ws = wb["Sheet3"]
    x.store_abbre = shop_abbre
    x.Shopify_App_API_Key = ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Key", ws["B"])[0].row)].value)
    x.Shopify_App_API_Password = ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Password", ws["B"])[0].row)].value)
    x.Shopify_App_API_Secret = ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Secret", ws["B"])[0].row)].value)
    x.Shopify_App_API_Url = ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Url", ws["B"])[0].row)].value)
    meta_tag = ws["C{}".format(sudby(lambda i:i.value=="Gsuite Account Meta Verification Tag", ws["B"])[0].row)].value)
    x.save()


    x = Shop()(shop_abbre)

    main_theme = sudby(lambda i:i.role == "main",x.shopify.Theme.find(status="any"))[0]
    asset_value = x.shopify.Asset.find(key="layout/theme.liquid",theme_id=main_theme.id).value
    new_asset_value = re.sub("<head>","<head>\n    %s"%(meta_tag),asset_value)
    asset = x.shopify.Asset.find(key="layout/theme.liquid",theme_id=main.id)
    asset.value = new_asset_value
    asset.save()
  n1()"""
  with open(os.path.expanduser("~/tavern/tavern/main.py"),"w") as f:
    f.write(x)
  os.system("python ~/tavern/tavern/main.py")
  os.system("rm ~/tavern/tavern/main.py")
  print("complete")
def part2():
  import os
  x = """import os
  os.chdir(os.path.expanduser("~/tavern/tavern"))
  from soda.can import *
  def n2():
    shop_abbre = OSA.log("input 3 letter shop abbreviation for shop")
    x = Shop()(shop_abbre)
    y = get_latest_download()
    Email_Address = OSA.log("Enter the personal email")
    a = "." + "|".join([Email_Address,"Client_Secret","GSuite"]) + ".json"
    shutil_move(cf, a)
    Binarydata().update_or_create(a)
    Google_Drive(Email_Address)
    Emailer(Email_Address)
  n2()"""
  with open(os.path.expanduser("~/tavern/tavern/main.py"),"w") as f:
    f.write(x)
  os.system("python ~/tavern/tavern/main.py")
  os.system("rm ~/tavern/tavern/main.py")
  print("complete")
def part3():
  import os
  x = """import os
  os.chdir(os.path.expanduser("~/tavern/tavern"))
  from soda.can import *
  def n3():
    shop_abbre = OSA.log("input 3 letter shop abbreviation for shop")
    x = Shop()(shop_abbre)
    y = get_latest_download()
    Email_Address = OSA.log("Enter the support email")
    a = "." + "|".join([Email_Address,"Client_Secret","GSuite"]) + ".json"
    shutil_move(cf, a)
    Binarydata().update_or_create(a)
    Google_Drive(Email_Address)
    Emailer(Email_Address)
  n3()"""
  with open(os.path.expanduser("~/tavern/tavern/main.py"),"w") as f:
    f.write(x)
  os.system("python ~/tavern/tavern/main.py")
  os.system("rm ~/tavern/tavern/main.py")
  print("complete")
def privacy_policy():
  import os
  x = """import os
  os.chdir(os.path.expanduser("~/tavern/tavern"))
  from soda.can import *
  def run():
    shop_abbre = OSA.log("input 3 letter shop abbreviation for shop")
    x = Shop()(shop_abbre)
    Facebook_Business_App_Privacy_Policy = 'Privacy Policy\n\n\nEffective date: __Effective_Date__\n\n\n__self.Business_Name__ ("us", "we", or "our") operates the http://__self.Domain_Name__/ website (the "Service").\n\nThis page informs you of our policies regarding the collection, use, and disclosure of personal data when you use our Service and the choices you have associated with that data. Our Privacy Policy for __self.Business_Name__ is managed through Free Privacy Policy.\n\nWe use your data to provide and improve the Service. By using the Service, you agree to the collection and use of information in accordance with this policy. Unless otherwise defined in this Privacy Policy, terms used in this Privacy Policy have the same meanings as in our Terms and Conditions, accessible from http://__self.Domain_Name__/\n\n\nInformation Collection And Use\n\nWe collect several different types of information for various purposes to provide and improve our Service to you.\n\nTypes of Data Collected\n\nPersonal Data\n\nWhile using our Service, we may ask you to provide us with certain personally identifiable information that can be used to contact or identify you ("Personal Data"). Personally identifiable information may include, but is not limited to:\n\n\nCookies and Usage Data\n\n\nUsage Data\n\nWe may also collect information how the Service is accessed and used ("Usage Data"). This Usage Data may include information such as your computer\'s Internet Protocol address (e.g. IP address), browser type, browser version, the pages of our Service that you visit, the time and date of your visit, the time spent on those pages, unique device identifiers and other diagnostic data.\n\nTracking & Cookies Data\nWe use cookies and similar tracking technologies to track the activity on our Service and hold certain information.\nCookies are files with small amount of data which may include an anonymous unique identifier. Cookies are sent to your browser from a website and stored on your device. Tracking technologies also used are beacons, tags, and scripts to collect and track information and to improve and analyze our Service.\nYou can instruct your browser to refuse all cookies or to indicate when a cookie is being sent. However, if you do not accept cookies, you may not be able to use some portions of our Service.\nExamples of Cookies we use:\n\n    Session Cookies. We use Session Cookies to operate our Service.\n    Preference Cookies. We use Preference Cookies to remember your preferences and various settings.\n    Security Cookies. We use Security Cookies for security purposes.\n\n\nUse of Data\n    \n__self.Business_Name__ uses the collected data for various purposes:    \n\n    To provide and maintain the Service\n    To notify you about changes to our Service\n    To allow you to participate in interactive features of our Service when you choose to do so\n    To provide customer care and support\n    To provide analysis or valuable information so that we can improve the Service\n    To monitor the usage of the Service\n    To detect, prevent and address technical issues\n\n\nTransfer Of Data\nYour information, including Personal Data, may be transferred to — and maintained on — computers located outside of your state, province, country or other governmental jurisdiction where the data protection laws may differ than those from your jurisdiction.\nIf you are located outside United States and choose to provide information to us, please note that we transfer the data, including Personal Data, to United States and process it there.\nYour consent to this Privacy Policy followed by your submission of such information represents your agreement to that transfer.\n__self.Business_Name__ will take all steps reasonably necessary to ensure that your data is treated securely and in accordance with this Privacy Policy and no transfer of your Personal Data will take place to an organization or a country unless there are adequate controls in place including the security of your data and other personal information.\n\nDisclosure Of Data\n\nLegal Requirements\n__self.Business_Name__ may disclose your Personal Data in the good faith belief that such action is necessary to:\n\n    To comply with a legal obligation\n    To protect and defend the rights or property of __self.Business_Name__\n    To prevent or investigate possible wrongdoing in connection with the Service\n    To protect the personal safety of users of the Service or the public\n    To protect against legal liability\n\n\nSecurity Of Data\nThe security of your data is important to us, but remember that no method of transmission over the Internet, or method of electronic storage is 100% secure. While we strive to use commercially acceptable means to protect your Personal Data, we cannot guarantee its absolute security.\n\nService Providers\nWe may employ third party companies and individuals to facilitate our Service ("Service Providers"), to provide the Service on our behalf, to perform Service-related services or to assist us in analyzing how our Service is used.\nThese third parties have access to your Personal Data only to perform these tasks on our behalf and are obligated not to disclose or use it for any other purpose.\n\n\n\nLinks To Other Sites\nOur Service may contain links to other sites that are not operated by us. If you click on a third party link, you will be directed to that third party\'s site. We strongly advise you to review the Privacy Policy of every site you visit.\nWe have no control over and assume no responsibility for the content, privacy policies or practices of any third party sites or services.\n\n\nChildren\'s Privacy\nOur Service does not address anyone under the age of 18 ("Children").\nWe do not knowingly collect personally identifiable information from anyone under the age of 18. If you are a parent or guardian and you are aware that your Children has provided us with Personal Data, please contact us. If we become aware that we have collected Personal Data from children without verification of parental consent, we take steps to remove that information from our servers.\n\n\nChanges To This Privacy Policy\nWe may update our Privacy Policy from time to time. We will notify you of any changes by posting the new Privacy Policy on this page.\nWe will let you know via email and/or a prominent notice on our Service, prior to the change becoming effective and update the "effective date" at the top of this Privacy Policy.\nYou are advised to review this Privacy Policy periodically for any changes. Changes to this Privacy Policy are effective when they are posted on this page.\n\n\nContact Us\nIf you have any questions about this Privacy Policy, please contact us:\n\n        By email: __self.Business_Email_Address__\n          \n                                        \n'
    Effective_Date = datetime.now().strftime("%B %d, %Y")
    import openpyxl

    wb = openpyxl.load_workbook(file_loc,data_only=True)
    ws = wb["Sheet3"]

    x.Business_Name = ws["C{}".format(sudby(lambda i:i.value=="Store name", ws["B"])[0].row)].value
    x.Domain_Name = ws["C{}".format(sudby(lambda i:i.value=="Primary Domain", ws["B"])[0].row)].value
    x.Business_Email_Address = ws["C{}".format(sudby(lambda i:i.value=="Customer email", ws["B"])[0].row)].value
    x.save()
    Facebook_Business_App_Privacy_Policy = Replacements(Facebook_Business_App_Privacy_Policy, "__self.Business_Name__", self.Business_Name, "__self.Domain_Name__", self.Domain_Name, "__self.Business_Email_Address__", self.Business_Email_Address, "__Effective_Date__", Effective_Date)
    text_to_docx(Facebook_Business_App_Privacy_Policy, "Facebook Business App Privacy Policy.docx")
    Privacy_Policy_URL = Google_Drive(Email_Address = Business_Email_Address).create(address = "Facebook Business App Privacy Policy.docx", public = True)
    rm("Facebook Business App Privacy Policy.docx")
    OSA.log("Save this Privacy Policy URL:\n%s"%(Privacy_Policy_URL))
  run()"""
  with open(os.path.expanduser("~/tavern/tavern/main.py"),"w") as f:
    f.write(x)
  os.system("python ~/tavern/tavern/main.py")
  os.system("rm ~/tavern/tavern/main.py")
  print("complete")
def final():
  import os
  x = """
  import os
  from soda.can import *
  def run():
    shop_abbre = OSA.log("input 3 letter shop abbreviation for shop")
    x = Shop()(shop_abbre)
    file_loc = OSA.log("input workbook location")
    import openpyxl

    wb = openpyxl.load_workbook(file_loc,data_only=True)
    ws = wb["Sheet3"]

    Update(x,Gmail_Email_Address =                        ws["C{}".format(sudby(lambda i:i.value=="Account email", ws["B"])[0].row)].value)
    Update(x,Business_Name =                              ws["C{}".format(sudby(lambda i:i.value=="Store name", ws["B"])[0].row)].value)
    Update(x,Country_Of_Business =                        ws["C{}".format(sudby(lambda i:i.value=="Country/Region", ws["B"])[0].row)].value)
    Update(x,First_Name =                                 ws["C{}".format(sudby(lambda i:i.value=="First name", ws["B"])[0].row)].value)
    Update(x,Last_Name =                                  ws["C{}".format(sudby(lambda i:i.value=="Last name", ws["B"])[0].row)].value)
    Update(x,Street_Address =                             ws["C{}".format(sudby(lambda i:i.value=="Street", ws["B"])[0].row)].value)
    Update(x,Street_Address_Line_2 =                      ws["C{}".format(sudby(lambda i:i.value=="Apartment, suit, etc (optional)", ws["B"])[0].row)].value)
    Update(x,State =                                      ws["C{}".format(sudby(lambda i:i.value=="State", ws["B"])[0].row)].value)
    Update(x,City =                                       ws["C{}".format(sudby(lambda i:i.value=="City", ws["B"])[0].row)].value)
    Update(x,ZIP_Code =                                   ws["C{}".format(sudby(lambda i:i.value=="Postal / ZIP Code", ws["B"])[0].row)].value)
    Update(x,Business_Phone_Number =                      Join("",re.findall("\d+",ws["C{}".format(sudby(lambda i:i.value=="Phone", ws["B"])[0].row)].value)))
    Update(x,Business_Email_Address =                     ws["C{}".format(sudby(lambda i:i.value=="Customer email", ws["B"])[0].row)].value)
    Update(x,Product_Return_Address =                     ws["C{}".format(sudby(lambda i:i.value=="Product Return Address", ws["B"])[0].row)].value)
    Update(x,Administrative_Url =                         ws["C{}".format(sudby(lambda i:i.value=="Administrative URL", ws["B"])[0].row)].value)
    Update(x,Facebook_Pixel_ID =                          ws["C{}".format(sudby(lambda i:i.value=="Facebook_Pixel_ID", ws["B"])[0].row)].value)
    Update(x,Shopify_App_API_Key =                        ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Key", ws["B"])[0].row)].value)
    Update(x,Shopify_App_API_Password =                   ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Password", ws["B"])[0].row)].value)
    Update(x,Shopify_App_API_Secret =                     ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Secret", ws["B"])[0].row)].value)
    Update(x,Shopify_App_API_Url =                        ws["C{}".format(sudby(lambda i:i.value=="Shopify_App_API_Url", ws["B"])[0].row)].value)
    Update(x,Domain_Name =                                ws["C{}".format(sudby(lambda i:i.value=="Primary Domain", ws["B"])[0].row)].value)
    Update(x,Facebook_Business_Ad_Account_ID =            ws["C{}".format(sudby(lambda i:i.value=="Facebook_Business_Ad_Account_ID", ws["B"])[0].row)].value)
    Update(x,Facebook_Business_App_Secret =               ws["C{}".format(sudby(lambda i:i.value=="Facebook_Business_App_Secret", ws["B"])[0].row)].value)
    Update(x,Facebook_Business_App_Token =                ws["C{}".format(sudby(lambda i:i.value=="Facebook_Business_App_Token", ws["B"])[0].row)].value)
    Update(x,AliExpress_Email =                           ws["C{}".format(sudby(lambda i:i.value=="AliExpress_Email", ws["B"])[0].row)].value)
    Update(x,AliExpress_Password = OSA.log("Please enter in your AliExpress Password. It will be used to log in to order items.",hidden=True))
    Update(x,Default_Product_Description = ws["C{}".format(sudby(lambda i:i.value=="Default_Product_Description", ws["B"])[0].row)].value)
    Update(x,AliExpress_Financial_Card_Information = {"Financial_Card_Number": OSA.display_dialog("Financial Card Number?",default_answer="",hidden=True),"Financial_Card_Expiration_Date": OSA.display_dialog("Financial Card Expiration Date [for example: 0125]?",default_answer="",hidden=True),"Financial_Card_CVV": OSA.display_dialog("Financial Card CVV?",default_answer="",hidden=True),"Financial_Card_Billing_First_Name": OSA.display_dialog("Financial Card Billing First Name?",default_answer="").title(),"Financial_Card_Billing_Last_Name": OSA.display_dialog("Financial Card Billing Last Name?",default_answer="").title(),"Financial_Card_Billing_Street_Address": OSA.display_dialog("Financial Card Billing Street Address?",default_answer=""),"Financial_Card_Billing_Street_Address_Line_2": OSA.display_dialog("Financial Card Billing Street Address Line 2 [enter in blank if this does not apply]?",default_answer=""),"Financial_Card_Billing_Country": OSA.display_dialog("Financial Card Billing Country?",default_answer=""),"Financial_Card_Billing_State": OSA.display_dialog("Financial Card Billing State [for example: New York]?",default_answer="").title(),"Financial_Card_Billing_City": OSA.display_dialog("Financial Card Billing City?",default_answer="").title(),"Financial_Card_Billing_ZIP_Code": Integer(OSA.display_dialog("Financial Card Billing ZIP Code?",default_answer="")),})
    x.Send_Tracking_Number = True
    x.save()
    Save(UniqueProductAlgorithm,shop=x.store_abbreviation,x='[A-Z][A-Z][A-Z][0-9][A-Z][A-Z][A-Z][A-Z][A-Z]')
  run()"""
  with open(os.path.expanduser("~/tavern/tavern/main.py"),"w") as f:
    f.write(x)
  os.system("python ~/tavern/tavern/main.py")
  os.system("rm ~/tavern/tavern/main.py")
  print("complete")



if __name__ == "__main__":
  import sys
  if sys.argv[1] == "part1":
    part1()
  if sys.argv[1] == "part2":
    part2()
  if sys.argv[1] == "part3":
    part3()
  if sys.argv[1] == "privacy_policy":
    privacy_policy()
  if sys.argv[1] == "final":
    final()
